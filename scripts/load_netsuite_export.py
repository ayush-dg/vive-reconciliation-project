"""
load_netsuite_export.py

Loads real NetSuite "Bill" rows (AsTech vendor) from every .xlsx file in
sample_data/netsuite_exports/ into demo_netsuite_bills, for the live-ERP-
matching demo (see migrations/010_add_demo_netsuite_bills.sql).

Looks inside sample_data/netsuite_exports/ rather than a hardcoded filename
so a new export dropped in that folder next month is picked up without a
code change. Reads the "<Vendor> Detail" sheet (the one with a payment_type
column), keeps only row_role == "detail" rows tagged payment_type == "Bill"
(excludes "Bill Payment" and "Bill Credit"), and drops group_header/page/
row_role/payment_type after filtering — only invoice_number, amount,
bill_date, source_file are loaded.

A file whose "Bill" rows can't be told apart from other rows (payment_type
column missing/shifted) is skipped with a warning rather than guessed at —
see the astech_matches_by_bill_jun28-jul1.xlsx case this was written
against, which lost its payment_type column entirely for ~92% of its detail
rows and was excluded from this demo's load by explicit decision.
"""

import glob
import os
import sys
from datetime import datetime, timezone

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

from src.lakehouse.connection import execute_query, get_connection

EXPORTS_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "sample_data",
    "netsuite_exports",
)

DETAIL_HEADER = ["file", "page", "group_header", "row_role", "payment_type", "date", "document_number", "amount"]
VALID_PAYMENT_TYPES = {"Bill", "Bill Payment", "Bill Credit"}


def _find_detail_sheet(workbook):
    for name in workbook.sheetnames:
        if "detail" in name.lower():
            return name
    return None


def _clean_amount(raw):
    s = str(raw).strip().replace("$", "").replace(",", "")
    negative = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    value = float(s)
    return -value if negative else value


def _clean_date(raw):
    s = str(raw).strip()
    return datetime.strptime(s, "%m/%d/%Y").date().isoformat()


def extract_bill_rows(xlsx_path):
    """Returns (bill_rows, counts) for one export file, or (None, None) if
    the file's payment_type column is unusable (see module docstring), or
    the file itself isn't a readable .xlsx (e.g. empty/placeholder)."""
    try:
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  SKIPPED {os.path.basename(xlsx_path)}: not a readable .xlsx file ({e})")
        return None, None
    sheet_name = _find_detail_sheet(workbook)
    if sheet_name is None:
        workbook.close()
        print(f"  SKIPPED {os.path.basename(xlsx_path)}: no '*Detail' sheet found")
        return None, None

    sheet = workbook[sheet_name]
    detail_rows = [
        dict(zip(DETAIL_HEADER, row))
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[3] == "detail"
    ]
    workbook.close()

    valid_tag_count = sum(1 for r in detail_rows if r["payment_type"] in VALID_PAYMENT_TYPES)
    if detail_rows and valid_tag_count / len(detail_rows) < 0.5:
        print(
            f"  SKIPPED {os.path.basename(xlsx_path)}: payment_type column looks "
            f"corrupted/shifted ({valid_tag_count}/{len(detail_rows)} rows have a "
            f"recognized payment_type) — excluded rather than guessed at"
        )
        return None, None

    counts = {"Bill": 0, "Bill Payment": 0, "Bill Credit": 0}
    bill_rows = []
    for r in detail_rows:
        pt = r["payment_type"]
        if pt in counts:
            counts[pt] += 1
        if pt == "Bill":
            bill_rows.append(
                {
                    "invoice_number": r["document_number"],
                    "amount": _clean_amount(r["amount"]),
                    "bill_date": _clean_date(r["date"]),
                    "source_file": r["file"],
                }
            )
    return bill_rows, counts


def main():
    xlsx_files = sorted(glob.glob(os.path.join(EXPORTS_DIR, "*.xlsx")))
    if not xlsx_files:
        print(f"No .xlsx files found in {EXPORTS_DIR}")
        return

    all_bill_rows = []
    print(f"Found {len(xlsx_files)} .xlsx file(s) in {EXPORTS_DIR}:")
    for path in xlsx_files:
        print(f" - {os.path.basename(path)}")
    print()

    for path in xlsx_files:
        bill_rows, counts = extract_bill_rows(path)
        if bill_rows is None:
            continue
        print(f"  {os.path.basename(path)}: Bill={counts['Bill']} Bill Payment={counts['Bill Payment']} Bill Credit={counts['Bill Credit']}")
        all_bill_rows.extend(bill_rows)

    print(f"\nTotal 'Bill' rows to load: {len(all_bill_rows)}")

    loaded_at = datetime.now(timezone.utc).isoformat()
    params = [
        (row["invoice_number"], row["amount"], row["bill_date"], row["source_file"], loaded_at)
        for row in all_bill_rows
    ]
    conn = get_connection()
    try:
        cursor = conn.cursor()
        if hasattr(cursor, "fast_executemany"):
            cursor.fast_executemany = True
        cursor.executemany(
            "INSERT INTO demo_netsuite_bills (invoice_number, amount, bill_date, source_file, loaded_at) "
            "VALUES (?, ?, ?, ?, ?)",
            params,
        )
        conn.commit()
    finally:
        conn.close()

    print(f"Loaded {len(all_bill_rows)} rows into demo_netsuite_bills (loaded_at={loaded_at})")

    count_result = execute_query("SELECT COUNT(*) AS n FROM demo_netsuite_bills")
    print(f"\nVerification — real row count in demo_netsuite_bills: {count_result[0]['n']}")

    sample = execute_query("SELECT invoice_number, amount, bill_date, source_file, loaded_at FROM demo_netsuite_bills ORDER BY id LIMIT 5")
    print("\nSample rows (queried back from the table):")
    for row in sample:
        print(f"  {row}")


if __name__ == "__main__":
    from dotenv import load_dotenv

    PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    load_dotenv(os.path.join(PROJECT_ROOT, ".env"))
    main()
